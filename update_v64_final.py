import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook("/sessions/nifty-festive-ramanujan/mnt/市場調査/Astromeda_ガントチャート_v64_商品整合性チェック版.xlsx")

# Add implementation status sheet
ws = wb.create_sheet("実装済み・検証結果", 0)

title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill('solid', fgColor='00C853')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1A1A2E')
normal_font = Font(name='Arial', size=10, color='FFFFFF')
normal_fill = PatternFill('solid', fgColor='16213E')
done_fill = PatternFill('solid', fgColor='00C853')
done_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin', color='333355'),
    right=Side(style='thin', color='333355'),
    top=Side(style='thin', color='333355'),
    bottom=Side(style='thin', color='333355')
)
wrap = Alignment(wrap_text=True, vertical='top')

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 45

# Title
ws.merge_cells('A1:F1')
c = ws['A1']
c.value = 'ASTROMEDA ECサイト — Phase 0 実装済み・検証結果レポート'
c.font = title_font; c.fill = title_fill; c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:F2')
c = ws['A2']
c.value = '実施日: 2026/04/07 | ビルド: ✅ SUCCESS (0 errors) | 検証: 全6ファイルPASS'
c.font = Font(name='Arial', size=10, color='00FF88'); c.fill = PatternFill('solid', fgColor='0D1117')
ws.row_dimensions[2].height = 25

# Headers
row = 4
headers = ['#', '修正内容', '対象ファイル', 'ステータス', 'ビルド', '検証詳細']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.border = thin_border
    c.alignment = Alignment(horizontal='center', vertical='center')

completed = [
    ['1', 'guides/beginners GPU表記\nRTX 4060→5060, 4070→5070, 4080→5080', 'app/routes/guides.beginners.tsx', '✅ 完了', '✅ PASS', 'フルHD: RTX 5060, WQHD: RTX 5070,\n4K: RTX 5080。全3箇所修正済み。'],
    ['2', 'guides/beginners CPU表記\nCore i5/i7→Core Ultra/Ryzen', 'app/routes/guides.beginners.tsx', '✅ 完了', '✅ PASS', 'Core Ultra 5/7 + Ryzen 5/7。\nNPU説明も追加。'],
    ['3', 'guides/beginners 予算別構成\n20万/30万/40万台', 'app/routes/guides.beginners.tsx', '✅ 完了', '✅ PASS', '全3ティア: RTX 5060/5070Ti/5090\n+ Ryzen 5/7/9 + DDR5。'],
    ['4', 'guides/beginners メモリ説明\nDDR4/DDR5→DDR5主流', 'app/routes/guides.beginners.tsx', '✅ 完了', '✅ PASS', '「DDR5が主流」「ASTROMEDA全モデル\nDDR5採用」に更新。'],
    ['5', 'guides/cospa エントリーGPU\nRTX 4060→5060', 'app/routes/guides.cospa.tsx', '✅ 完了', '✅ PASS', 'GPU: RTX 5060\nCPU: Ryzen 5 / Core Ultra 5'],
    ['6', 'guides/cospa ミドルハイGPU\nRTX 4070 SUPER→5070/5070Ti', 'app/routes/guides.cospa.tsx', '✅ 完了', '✅ PASS', 'GPU: RTX 5070 / RTX 5070Ti\nCPU: Ryzen 7 / Core Ultra 7'],
    ['7', 'guides/cospa フラッグシップGPU\nRTX 4080 SUPER/4090→5080/5090', 'app/routes/guides.cospa.tsx', '✅ 完了', '✅ PASS', 'GPU: RTX 5080 / RTX 5090\nCPU: Ryzen 9 / Core Ultra 9'],
    ['8', 'guides/cospa 失敗パターン例\nRTX 4090→RTX 5090', 'app/routes/guides.cospa.tsx', '✅ 完了', '✅ PASS', '「RTX 5090にローエンドCPUを\n合わせる」に修正。'],
    ['9', 'guides/streaming OBS推奨\nRTX 4060→5060', 'app/routes/guides.streaming.tsx', '✅ 完了', '✅ PASS', '推奨: RTX 5060以上。\nXSplit: RTX 5070以上。'],
    ['10', 'guides/streaming NVENC説明\nRTX 40→50シリーズ', 'app/routes/guides.streaming.tsx', '✅ 完了', '✅ PASS', '「RTX 50シリーズのNVENC」に統一。\nAV1も50シリーズ対応に。'],
    ['11', 'guides/streaming 1PC配信スペック\nRTX 4070→5070', 'app/routes/guides.streaming.tsx', '✅ 完了', '✅ PASS', 'RTX 5070以上 + Ryzen 7 /\nCore Ultra 7以上。'],
    ['12', 'guides/streaming 推奨スペック表\n全3ティア更新', 'app/routes/guides.streaming.tsx', '✅ 完了', '✅ PASS', 'ライト: RTX 5060\nスタンダード: RTX 5070Ti\nハイエンド: RTX 5080'],
    ['13', 'コラボ数統一\n23/25/26→「23タイトル以上」', 'astromeda-data.ts\n_index.tsx', '✅ 完了', '✅ PASS', 'MARQUEE, meta description,\nOG description 全て統一。'],
    ['14', 'CollabGrid null安全性\n!→?.', 'CollabGrid.tsx', '✅ 完了', '✅ PASS', 'shopifyCol!.image!.url →\nshopifyCol?.image?.url ?? \'\''],
    ['15', 'RTX 4000番台 完全除去\napp/ディレクトリ全体', 'app/ 全ファイル', '✅ 完了', '✅ PASS', 'grep確認: 0件。\n4060/4070/4080/4090 全てゼロ。'],
]

for i, d in enumerate(completed):
    r = row + 1 + i
    for j, val in enumerate(d, 1):
        c = ws.cell(row=r, column=j, value=val)
        c.font = normal_font; c.fill = normal_fill; c.border = thin_border
        c.alignment = wrap
    ws[f'D{r}'].font = done_font; ws[f'D{r}'].fill = done_fill
    ws[f'E{r}'].font = done_font; ws[f'E{r}'].fill = done_fill
    ws.row_dimensions[r].height = 50

# Audit findings section
r = row + len(completed) + 2
ws.merge_cells(f'A{r}:F{r}')
c = ws[f'A{r}']
c.value = '■ システム全体監査で発見した追加課題（Phase 0で修正済み + 今後の課題）'
c.font = Font(name='Arial', size=13, bold=True, color='00F0FF')
c.fill = PatternFill('solid', fgColor='0D1117')
ws.row_dimensions[r].height = 30

r += 1
audit_headers = ['#', '課題', '重要度', 'ステータス', '層', '説明']
for i, h in enumerate(audit_headers, 1):
    c = ws.cell(row=r, column=i, value=h)
    c.font = header_font; c.fill = header_fill; c.border = thin_border

audit_items = [
    ['1', 'GPU/CPU表記不整合', '致命的', '✅ 修正済み', 'DNA', 'ガイド3ページ全修正完了'],
    ['2', 'コラボ数不整合(23/25/26)', '高', '✅ 修正済み', 'DNA', '「23タイトル以上」に統一'],
    ['3', 'CollabGrid null assertion', '高', '✅ 修正済み', '臓器', 'オプショナルチェーンに変更'],
    ['4', 'PC_COLORS GraphQL手動定義', '中', '⏳ 今後対応', 'DNA', 'IP_HANDLESと同様の自動生成に'],
    ['5', 'API失敗時のサイレント劣化', '中', '⏳ 今後対応', '神経系', 'ユーザー通知を追加'],
    ['6', 'カテゴリ画像ハードコード', '中', '⏳ 今後対応', '神経系', 'Shopify API動的取得に変更'],
    ['7', 'PC_TIERS価格ハードコード', '中', '⏳ 今後対応', 'DNA', 'Shopify実データ連動に'],
    ['8', 'エラー監視なし(本番)', '中', '⏳ 今後対応', '免疫系', 'Sentry等のエラー追跡導入'],
    ['9', 'setupページ画像未設定', '致命的', '⏳ Shopify側', '臓器', 'Shopifyに8色styleページ作成必要'],
    ['10', 'Rate limit メモリリーク', '低', '⏳ 今後対応', '免疫系', '大規模時に要対策'],
]

for i, d in enumerate(audit_items):
    r2 = r + 1 + i
    for j, val in enumerate(d, 1):
        c = ws.cell(row=r2, column=j, value=val)
        c.font = normal_font; c.fill = normal_fill; c.border = thin_border
        c.alignment = wrap
    if '修正済み' in d[3]:
        ws.cell(row=r2, column=4).font = done_font
        ws.cell(row=r2, column=4).fill = done_fill
    elif 'Shopify' in d[3]:
        ws.cell(row=r2, column=4).font = Font(name='Arial', size=10, bold=True, color='000000')
        ws.cell(row=r2, column=4).fill = PatternFill('solid', fgColor='FFD700')
    ws.row_dimensions[r2].height = 35

wb.save("/sessions/nifty-festive-ramanujan/mnt/市場調査/Astromeda_ガントチャート_v64_商品整合性チェック版.xlsx")
print("v64 updated with implementation results")
