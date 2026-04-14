import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

COLORS = {
    'header_bg': '0F3460', 'dark_bg': '1A1A2E', 'white': 'FFFFFF',
    'complete': 'C8E6C9', 'partial': 'FFF9C4', 'stub': 'FFE0B2', 'missing': 'FFCDD2',
    'phase0': 'E8D5F5', 'phase1': 'D4E6F1', 'phase2': 'FADBD8', 'phase3': 'D5F5E3',
    'phase4': 'FCF3CF', 'phase5': 'D6EAF8', 'phase6': 'F9E79F', 'phase7': 'ABEBC6',
    'phase8': 'F5B7B1', 'phase9': 'D2B4DE', 'phase10': 'AED6F1', 'phase11': 'A9DFBF',
}

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
)
header_font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
header_fill = PatternFill('solid', fgColor=COLORS['header_bg'])
normal_font = Font(name='Arial', size=8, color='333333')
phase_font = Font(name='Arial', bold=True, size=10, color=COLORS['dark_bg'])
critical_font = Font(name='Arial', bold=True, size=8, color='FF0000')
warn_font = Font(name='Arial', bold=True, size=8, color='FF8F00')

# ========== TASK DATA ==========
# (ID, Task, ExistingFile, Status, ExactWork, AcceptanceCriteria, TestCommand, Days, Deps, Priority, Phase)

tasks = [
    # ===== Phase 0: 受精・着床 =====
    ('PH0', '【Phase 0】受精・着床 — 環境構築・外部サービス接続', '', '', '', '', '', '', [], '', 'phase0'),

    ('0-01', 'Anthropic APIキー取得・.env設定', '.env (既存)', '要確認',
     'ANTHROPIC_API_KEY を .env に設定。ClaudeProvider で接続テスト実行',
     'npm run test -- agents/providers でClaudeProvider初期化成功', 'npm run test', '0.5', [], 'CRITICAL', 'phase0'),

    ('0-02', 'Google Gemini APIキー取得・.env設定', '.env (既存)', '要確認',
     'GOOGLE_GEMINI_API_KEY を .env に設定。GeminiProvider で接続テスト実行',
     'npm run test -- agents/providers でGeminiProvider初期化成功', 'npm run test', '0.5', [], 'CRITICAL', 'phase0'),

    ('0-03', 'Google Cloud Service Account JSON作成', 'なし', 'MISSING',
     'GA4 Data API + Search Console API 用のサービスアカウントを作成。JSONキーをダウンロードし GOOGLE_SERVICE_ACCOUNT_JSON として .env に設定',
     'agents/data-collection/ga4-client.ts の getAccessToken() がJWT署名付きトークンを返す', 'npm run test -- agents/data-collection', '1', [], 'CRITICAL', 'phase0'),

    ('0-04', 'Fly.ioアカウント作成・CLIインストール', 'なし', 'MISSING',
     'powershell: iwr https://fly.io/install.ps1 -useb | iex でCLIインストール。fly auth login でログイン。fly apps create astromeda-agents',
     'fly status が正常応答。fly apps list に astromeda-agents が表示', 'fly status', '0.5', [], 'CRITICAL', 'phase0'),

    ('0-05', 'Fly.io PostgreSQL プロビジョニング', 'なし', 'MISSING',
     'fly postgres create --name astromeda-db --region nrt。fly postgres attach astromeda-db --app astromeda-agents',
     'fly postgres connect -a astromeda-db で接続成功。\\dt で空テーブル一覧', 'fly postgres connect -a astromeda-db', '0.5', ['0-04'], 'CRITICAL', 'phase0'),

    ('0-06', 'PostgreSQL 6テーブルスキーマ作成', 'agents/data-collection/data-models.ts (型定義あり)', 'PARTIAL',
     'data-models.ts の型定義を元にSQLマイグレーション作成: analytics_daily, search_console_daily, ai_visibility_weekly, competitor_weekly, feedback_history, approval_queue。lib/databases/schema.sql + migrate.ts を作成',
     '6テーブルが作成され、INSERT/SELECT が正常動作。マイグレーションのロールバックも動作確認',
     'npx tsx lib/databases/migrate.ts && npm run test -- lib/databases', '1', ['0-05'], 'CRITICAL', 'phase0'),

    ('0-07', 'Slack 4チャンネル作成・Bot設定', 'なし', 'MISSING',
     'Slack App作成（OAuth: chat:write, channels:read, reactions:write）。#astromeda-通知, #astromeda-品質, #astromeda-障害, #astromeda-情報 の4チャンネル作成。SLACK_BOT_TOKEN, SLACK_SIGNING_SECRET を .env に設定',
     '各チャンネルにテストメッセージが送信されること。agents/core/agent-bus.ts のSlack通知hookが動作',
     'npm run test -- agents/approval', '1', [], 'HIGH', 'phase0'),

    ('0-08', 'GA4 JWT認証実装', 'agents/data-collection/ga4-client.ts (TODO: Phase 2-G完成時に実装)', 'STUB',
     'ga4-client.ts の getAccessToken() メソッドを修正。google-auth-library の JWT クラスを使用し、サービスアカウントJSONからアクセストークンを取得する実装に差し替え。スタブパスを残しつつ isConnected=true 時に実APIを使用',
     'GA4 Data API (analyticsdata.googleapis.com/v1beta) から実データ取得。getDailySummary() が本物のPV/セッション数を返す',
     'npm run test -- agents/data-collection/ga4', '1', ['0-03'], 'HIGH', 'phase0'),

    ('0-09', 'GSC JWT認証実装', 'agents/data-collection/gsc-client.ts (TODO: 同上)', 'STUB',
     'gsc-client.ts の getAccessToken() を GA4と同様に修正。Search Console API (searchconsole.googleapis.com/v3) への実接続',
     'getTopQueries() が実際の検索クエリデータを返す。detectRankingChanges() が実データで動作',
     'npm run test -- agents/data-collection/gsc', '1', ['0-03'], 'HIGH', 'phase0'),

    # ===== Phase 1: 胚形成 — AIルーティング強化 =====
    ('PH1', '【Phase 1】胚形成 — デュアルAI基盤の完成', '', '', '', '', '', '', [], '', 'phase1'),

    ('1-01', 'AIProvider統一インターフェース検証', 'agents/core/ai-brain.ts (343行 完全実装)', 'COMPLETE',
     '既存ai-brain.tsのAIBrain classがClaude API接続を実装済み。Tier A-D割当テーブル（設計書v7の47体×モデル対応表）をconfigとして追加。AIRouter classを新規作成: agents/core/ai-router.ts',
     'AIRouter.getModel(agentId) が設計書通りのTier (A=Sonnet, B=Haiku, C=Flash, D=Flash-Lite) を返す。全47エージェントIDに対してnullを返さない',
     'npm run test -- agents/core/ai-router', '1', ['0-01', '0-02'], 'CRITICAL', 'phase1'),

    ('1-02', 'GeminiProvider実装', 'agents/core/ai-brain.ts (ClaudeのみでGemini未実装)', 'MISSING',
     'agents/core/gemini-provider.ts を新規作成。@google/generative-ai パッケージでGemini Flash / Flash-Lite 接続。AIBrainと同じインターフェース (generateText, generateJSON, generateImage) を実装',
     'GeminiProvider.generateText("テスト") がGemini Flashから応答を返す。Imagen APIで画像生成が動作',
     'npm run test -- agents/core/gemini-provider', '2', ['0-02'], 'CRITICAL', 'phase1'),

    ('1-03', 'FallbackManager実装', 'agents/core/circuit-breaker.ts (227行 完全実装)', 'PARTIAL',
     '既存circuit-breaker.tsはopen/half-open/closed状態管理あり。これを拡張してFallbackManager class (agents/core/fallback-manager.ts) を新規作成: 3回リトライ(指数バックオフ) → Tier別代替AI切替 → Slack#astromeda-障害通知 → 5分復旧チェック',
     'Claude API停止シミュレーションで自動的にGeminiに切替。復旧時に自動復帰。Slack通知が届く。全プロセスが5秒以内',
     'npm run test -- agents/core/fallback-manager', '1.5', ['1-01', '1-02'], 'CRITICAL', 'phase1'),

    ('1-04', 'Agent Bus Slack統合', 'agents/core/agent-bus.ts (307行 完全実装)', 'COMPLETE',
     '既存agent-bus.tsにsecurity hook + feedback hook あり。Slack WebhookをNotificationBus (agents/core/notification-bus.ts 新規) として実装。Critical→即時, High→1h, Medium→日次, Low→週次 のレベル別配信',
     '4チャンネルにそれぞれ適切なレベルのメッセージが配信される。CRITICALは10秒以内に通知',
     'npm run test -- agents/core/notification-bus', '1.5', ['0-07'], 'HIGH', 'phase1'),

    # ===== Phase 2: 心臓形成 — Fly.ioデプロイ =====
    ('PH2', '【Phase 2】心臓形成 — 最初のデプロイ', '', '', '', '', '', '', [], '', 'phase2'),

    ('2-01', 'Fly.io Dockerfile/fly.toml作成', 'なし', 'MISSING',
     'プロジェクトルートに fly.toml 作成 (app=astromeda-agents, region=nrt, vm.size=shared-cpu-1x, memory=256MB)。Dockerfile作成 (Node 22 alpine, npm ci --production, agents/ + lib/ のみコピー)',
     'fly deploy --local-only でビルド成功。fly status でrunning。/api/health エンドポイントが200返却',
     'fly deploy && fly status', '1', ['0-04', '0-06'], 'CRITICAL', 'phase2'),

    ('2-02', 'Commander + Registry + Bus 初回デプロイ', 'agents/l0/commander.ts (343行), agents/registry/agent-registry.ts (155行)', 'COMPLETE',
     'fly.toml のentrypointをagents/registration/agent-registration.ts に設定。環境変数をfly secrets set。最小構成 (Commander + Registry + AgentBus + 5 TeamLeads) でデプロイ',
     'fly logs にCommander起動ログ。/api/health で {"status":"healthy","agents":7}。Slack #astromeda-通知 に起動通知',
     'fly deploy && fly logs --app astromeda-agents', '1', ['2-01', '1-03'], 'CRITICAL', 'phase2'),

    ('2-03', 'ヘルスチェック + 死活監視設定', 'agents/core/health-monitor.ts (既存)', 'PARTIAL',
     'fly.toml に [[services.http_checks]] url="/api/health" interval="30s" timeout="5s" 追加。health-monitor.ts にFly.io Machines restart hook追加。3回連続失敗でSlack #astromeda-障害 通知',
     'fly machine stop → 自動再起動。ヘルスチェック失敗時にSlack通知。fly dashboard で緑ステータス',
     'fly machine stop $(fly machine list -q) && sleep 60 && fly status', '1', ['2-02'], 'HIGH', 'phase2'),

    # ===== Phase 3: 神経系発達 — L1+L2基盤チーム =====
    ('PH3', '【Phase 3】神経系発達 — チームリード + インフラチーム', '', '', '', '', '', '', [], '', 'phase3'),

    ('3-01', 'L1チームリード名称整合', 'agents/l1/ (data-lead, engineering-lead, marketing-lead, product-lead, sales-lead)', 'PARTIAL',
     '既存5リードは一般名称。設計書v7の6チーム (集客/転換/LTV/基盤/情報/司令塔) と対応付け: marketing-lead→acquisition-lead, sales-lead→conversion-lead, product-lead→ltv-lead, engineering-lead→infrastructure-lead, data-lead→intelligence-lead。ファイルをリネームorエイリアス作成',
     '全L1リードが設計書v7のチーム名で登録される。agent-registration.ts のインポートパスが正常動作',
     'npm run test -- agents/l1 && npm run test -- agents/registration', '1', ['2-02'], 'HIGH', 'phase3'),

    ('3-02', 'Error Monitor強化', 'agents/l2/error-monitor.ts (既存)', 'PARTIAL',
     '既存error-monitor.tsにSlack #astromeda-障害 への自動通知を追加。エラーパターン学習 (5分以内に同一エラー3回→CRITICAL昇格) 実装。Fly.io logs連携',
     'アプリケーションエラー発生時に30秒以内にSlack通知。エラー集計ダッシュボードAPI (/api/errors/summary) が動作',
     'npm run test -- agents/l2/error-monitor', '1.5', ['1-04', '2-02'], 'CRITICAL', 'phase3'),

    ('3-03', 'P05 Deploy Pipeline Fly.io対応', 'agents/pipelines/pipeline-definitions.ts (P10-P12にdeploy系あり)', 'PARTIAL',
     '既存P10(deploy)をFly.io Machines API対応に更新。fly deploy → ヘルスチェック → Slack通知 → ロールバック機構。GitHub Actions workflow (.github/workflows/deploy-agents.yml) 作成',
     'git push で自動デプロイ → Fly.io更新 → ヘルスチェック成功 → Slack通知。ロールバックコマンドが動作',
     'npm run test -- agents/pipelines/pipeline-engine', '2', ['2-03'], 'HIGH', 'phase3'),

    # ===== Phase 4: 基盤臓器 — データ収集実稼働 =====
    ('PH4', '【Phase 4】基盤臓器形成 — データ収集の実稼働', '', '', '', '', '', '', [], '', 'phase4'),

    ('4-01', 'G-1 GA4日次収集バッチ実装', 'agents/data-collection/ga4-client.ts (507行 40%実装)', 'PARTIAL',
     'batchCollect() は実装済み。cron-trigger (毎朝3:00 JST) を agents/core/scheduler.ts に追加。結果を PostgreSQL analytics_daily テーブルにINSERT。Slack #astromeda-通知 に日次サマリー送信',
     '毎朝3:00にGA4からデータ取得→DB保存→Slack通知の一連が動作。analytics_daily テーブルに日付別レコード',
     'npm run test -- agents/data-collection/ga4 && npx tsx agents/data-collection/ga4-client.ts --run-now', '1.5', ['0-08', '0-06'], 'HIGH', 'phase4'),

    ('4-02', 'G-2 GSC日次収集バッチ実装', 'agents/data-collection/gsc-client.ts (386行 50%実装)', 'PARTIAL',
     '同上パターン。3日遅延データを取得。search_console_daily テーブルにINSERT。detectRankingChanges() で順位変動検知→Slack #astromeda-品質 通知',
     '毎朝実行でGSCデータ取得→DB保存。順位10位以上下落時にSlack通知',
     'npm run test -- agents/data-collection/gsc && npx tsx agents/data-collection/gsc-client.ts --run-now', '1.5', ['0-09', '0-06'], 'HIGH', 'phase4'),

    ('4-03', 'G-3 AI可視性週次チェック実装', 'agents/data-collection/ai-visibility-checker.ts (281行 Stubのみ)', 'STUB',
     'checkVisibility() をPerplexity API / ChatGPT Search / Google AI Overview 実装に差し替え。20-30クエリを各プラットフォームに送信。結果を ai_visibility_weekly テーブルにINSERT。週次レポート→Slack #astromeda-情報',
     '月曜6:00にAI検索エンジンを実クエリ→結果DB保存。"Astromeda"の言及有無・順位・文脈が記録される',
     'npm run test -- agents/data-collection/ai-visibility', '2', ['0-06'], 'HIGH', 'phase4'),

    ('4-04', 'G-4 競合監視週次チェック実装', 'agents/data-collection/competitor-scraper.ts (412行 Stubのみ)', 'STUB',
     'runWeeklyPCCheck() を実スクレイピング実装に差し替え。Cheerio (HTML解析) で7メーカー公式サイトの商品ページから価格・スペック取得。Amazon PA-API 5.0 でガジェット価格取得。competitor_weekly テーブルにINSERT',
     '水曜5:00に7メーカー+Amazon/楽天のデータ取得→DB保存。価格変動のある競合モデルがSlack #astromeda-情報 に通知',
     'npm run test -- agents/data-collection/competitor', '2', ['0-06'], 'HIGH', 'phase4'),

    # ===== Phase 5: 成長臓器I — SNSプロバイダー実装 =====
    ('PH5', '【Phase 5】成長臓器I — 外部API接続（SNS/広告）', '', '', '', '', '', '', [], '', 'phase5'),

    ('5-01', 'XProvider実装 (X API v2)', 'agents/providers/sns-providers.ts (XTwitterProvider TODO)', 'STUB',
     'XTwitterProvider の executeRealPost() を X API v2 (OAuth 2.0 PKCE) で実装。postTweet(text, mediaIds?) + getEngagement(tweetId) + uploadMedia(buffer)。環境変数: X_CLIENT_ID, X_CLIENT_SECRET',
     'テスト投稿がXに表示される（テストアカウントで確認）。エンゲージメントデータが取得できる',
     'npm run test -- agents/providers/sns', '1.5', ['0-07'], 'MEDIUM', 'phase5'),

    ('5-02', 'InstagramProvider実装 (Meta Graph API)', 'agents/providers/sns-providers.ts (InstagramProvider TODO)', 'STUB',
     'InstagramProvider の executeRealPost() を Meta Graph API で実装。publishPost(imageUrl, caption) + publishCarousel(items) + getInsights(mediaId)。環境変数: META_APP_ID, META_APP_SECRET, IG_BUSINESS_ID',
     'テスト画像投稿がInstagramに表示。インサイトデータが取得できる',
     'npm run test -- agents/providers/sns', '1.5', ['0-07'], 'MEDIUM', 'phase5'),

    ('5-03', 'TikTokProvider実装 (Content Posting API)', 'agents/providers/sns-providers.ts (TikTokProvider TODO)', 'STUB',
     'TikTokProvider の executeRealPost() を TikTok Content Posting API (OAuth 2.0) で実装。publishVideo(videoUrl) + getPerformance(videoId)。環境変数: TIKTOK_CLIENT_KEY, TIKTOK_CLIENT_SECRET',
     'テスト動画投稿がTikTokに表示。パフォーマンスデータが取得できる',
     'npm run test -- agents/providers/sns', '1.5', ['0-07'], 'MEDIUM', 'phase5'),

    ('5-04', 'GoogleAdsProvider実装 (Google Ads API v17)', 'agents/providers/ads-providers.ts (GoogleAdsProvider TODO)', 'STUB',
     'GoogleAdsProvider の getCampaigns()/getPerformance() を Google Ads API v17 で実装。OAuth 2.0 + Developer Token + MCC。環境変数: GOOGLE_ADS_DEV_TOKEN, GOOGLE_ADS_MCC_ID',
     'キャンペーン一覧が取得できる。ROAS/CTR/CPCの実データが返る',
     'npm run test -- agents/providers/ads', '1.5', [], 'MEDIUM', 'phase5'),

    ('5-05', 'MetaAdsProvider + LINEProvider実装', 'agents/providers/ads-providers.ts (TODO)', 'STUB',
     'MetaAdsProvider: Meta Marketing API。LINEAdsProvider: LINE Messaging API + LINE Ads API。環境変数: META_AD_ACCOUNT_ID, LINE_CHANNEL_ID, LINE_CHANNEL_SECRET',
     '各広告プラットフォームからキャンペーンデータ取得。LINE Messaging APIでテストメッセージ送信成功',
     'npm run test -- agents/providers/ads', '2', [], 'MEDIUM', 'phase5'),

    # ===== Phase 6: 成長臓器II — 承認フロー実稼働 =====
    ('PH6', '【Phase 6】成長臓器II — 承認フロー + Slack連携', '', '', '', '', '', '', [], '', 'phase6'),

    ('6-01', 'SlackApprovalBot実装 (Block Kit UI)', 'agents/approval/approval-orchestrator.ts (410行 完全実装)', 'PARTIAL',
     'ApprovalOrchestrator は完全実装だがSlack Block Kit UI未実装。agents/approval/slack-approval-bot.ts を新規作成: ✅承認/❌却下/✏️修正依頼ボタン。Interactive Messages API (action endpoint)。approval-orchestrator.ts の manualApproval/manualRejection と接続',
     'Slack #astromeda-通知 に承認リクエストが Block Kit UI で表示。ボタン押下で承認/却下がDB反映',
     'npm run test -- agents/approval', '2', ['0-07'], 'HIGH', 'phase6'),

    ('6-02', 'I-1〜I-6 承認フロー接続', 'agents/approval/index.ts (export済み)', 'PARTIAL',
     '各コンテンツ種別 (SNS/広告/バナー/メール/記事/価格変更) の承認フローを SlackApprovalBot + ContentApprovalGate (新規) 経由で接続。agents/approval/content-approval-gate.ts: 承認済みコンテンツのみProvider.execute()を許可',
     '未承認コンテンツがProviderに到達しない。承認後に自動投稿/入稿される。feedback_history テーブルに記録される',
     'npm run test -- agents/approval', '2', ['6-01', '5-01'], 'HIGH', 'phase6'),

    ('6-03', 'FeedbackAnalyzer学習サイクル接続', 'agents/approval/feedback-analyzer.ts (387行 完全実装)', 'COMPLETE',
     'FeedbackAnalyzer は完全実装。Self-Improvement (#33) との接続を強化: runLearningCycle() の結果をSelf-Improvement.optimizePrompt() に渡す。月次レポートを Slack #astromeda-品質 に自動送信',
     '承認/却下パターンが分析され、承認率トレンドがSlackに月次レポートされる。3ヶ月で50%→80%の目標に向けた改善提案が生成される',
     'npm run test -- agents/approval/feedback-analyzer', '1', ['6-02'], 'HIGH', 'phase6'),

    # ===== Phase 7: 感覚器 — 情報チーム完全サイクル =====
    ('PH7', '【Phase 7】感覚器発達 — 情報チーム6ステップサイクル', '', '', '', '', '', '', [], '', 'phase7'),

    ('7-01', 'lib/validation/types.ts + statistical-engine.ts', 'なし', 'MISSING',
     'agents/lib/validation/types.ts: TestScenario, AttackPayload, Vulnerability, GoNoGoReport, RoundResult, VulnMap 型定義。agents/lib/validation/statistical-engine.ts: tTest(a,b), cohenD(a,b), confidenceInterval(data,α), cusum(series) の4関数。jstat ライブラリ使用',
     'tTest([1,2,3],[4,5,6]) が {p:number, significant:boolean} を返す。cohenD が EffectSize を返す。95% CI が正しい区間',
     'npm run test -- agents/lib/validation/statistical-engine', '2', [], 'CRITICAL', 'phase7'),

    ('7-02', 'lib/validation/round-executor.ts', 'なし', 'MISSING',
     'executeRound(config: RoundConfig): RoundResult — N回反復実行+統計集計。checkConvergence(results): boolean — CV≤15%判定。aggregateResults(rounds): Summary — 全ラウンド統合',
     'Round1(N=10)でCV計算正常。CV>30%で追加10回実行。Round2(N=20)で5パラメータ変動テスト。全結果がRoundResult型で返る',
     'npm run test -- agents/lib/validation/round-executor', '1.5', ['7-01'], 'CRITICAL', 'phase7'),

    ('7-03', 'lib/validation/sandbox-manager.ts', 'なし', 'MISSING',
     'create(config): Sandbox — Fly.io Machines APIでサンドボックスVM作成。destroy(id): void — VM破棄。sanitizeData(src): DataSet — テストデータ生成。takeBaseline(id): Metrics — ベースライン計測',
     'Fly.io上にサンドボックスVM作成→テスト実行→VM破棄のライフサイクルが動作。リソースリークなし',
     'npm run test -- agents/lib/validation/sandbox-manager', '2', ['0-04'], 'HIGH', 'phase7'),

    ('7-04', 'lib/validation/attack-engine.ts + vulnerability-mapper.ts', 'なし', 'MISSING',
     'attack-engine.ts: executeAttack(plan), mutatePayloads(p), varyConditions(c), repeatAttack(config)。vulnerability-mapper.ts: buildVulnMap(vulns), scoreCVSS(vuln), generatePatch(vuln), createRemediationPlan(map)',
     'Layer1攻撃 (10回反復×20突然変異) が実行可能。CVSS v4.0スコアリングが正確。パッチ自動生成が動作',
     'npm run test -- agents/lib/validation/attack-engine && npm run test -- agents/lib/validation/vulnerability-mapper', '3', ['7-01'], 'CRITICAL', 'phase7'),

    ('7-05', 'AI Capability Validator (#39) 実装', 'なし (設計書v7にTypeScript仕様あり)', 'MISSING',
     'agents/l2/ai-capability-validator.ts: receiveReport→createSandbox→executeRound1→executeRound2→runStatisticalTests→generateReport→requestApproval の7ステップ。入力型: ValidateRequest, SandboxRequest, RoundConfig, JudgmentInput。出力型: ValidationPlan, SandboxInstance, RoundResult, GoNoGoReport',
     'HIGH評価のテクノロジーレポートを入力→サンドボックス作成→2ラウンド検証→統計判定(t検定p<0.05, Cohen d≥0.5)→Go/No-Go報告書生成',
     'npm run test -- agents/l2/ai-capability-validator', '3', ['7-02', '7-03'], 'CRITICAL', 'phase7'),

    ('7-06', 'AI Security Auditor (#40) 実装', 'agents/l2/security-agent.ts (既存 名称不一致)', 'PARTIAL',
     '既存security-agent.tsを拡張/リネーム→agents/l2/ai-security-auditor.ts: analyzeThreat→executeAttack→executeRepeated→buildVulnMap→generatePatches→createRemediationPlan の6ステップ。Layer1反復攻撃(10回×20変異×4時間帯×4負荷)',
     'CRITICALセキュリティ脅威→攻撃実行→脆弱性マップ(CVSS v4.0)→プロンプトパッチ自動生成→改善計画の一連が動作',
     'npm run test -- agents/l2/ai-security-auditor', '3', ['7-04'], 'CRITICAL', 'phase7'),

    ('7-07', 'P15 Proof Validation Pipeline (7ステップ)', 'agents/pipelines/pipeline-definitions.ts に追加', 'MISSING',
     'pipeline-definitions.ts にP15定義追加: trigger=manual|tech-radar-high, steps=[analyze,sandbox,experiment,round1,round2,judgment,deploy-support], timeout=72h。pipeline-engine.ts で実行可能',
     'P15パイプラインを手動トリガー→7ステップが順序通り実行→GoNoGoReportが生成→Slack #astromeda-情報 に通知',
     'npm run test -- agents/pipelines', '1.5', ['7-05'], 'CRITICAL', 'phase7'),

    ('7-08', 'P16 Red Team Diagnosis Pipeline (8ステップ)', 'agents/pipelines/pipeline-definitions.ts に追加', 'MISSING',
     'P16定義追加: trigger=security-sentinel-critical|monthly, steps=[analyze,build-env,attack,repeat,vuln-map,defense-gen,verify,deploy-fix], timeout=48h',
     'P16パイプラインをトリガー→8ステップ実行→脆弱性マップ+パッチ生成→Slack #astromeda-障害 に通知',
     'npm run test -- agents/pipelines', '1.5', ['7-06'], 'CRITICAL', 'phase7'),

    # ===== Phase 8: 高次脳 — Shopify統合 + 全体結合 =====
    ('PH8', '【Phase 8】高次脳機能 — EC統合 + 全エージェント結合', '', '', '', '', '', '', [], '', 'phase8'),

    ('8-01', 'Shopify Webhook統合', 'agents/core/shopify-admin.ts (既存), agents/integration/ (既存)', 'PARTIAL',
     '既存shopify-admin.tsにAdmin API接続あり。Webhook受信エンドポイント追加: orders/create, products/update, inventory_levels/update, carts/update。各Webhookを対応エージェント (Product Catalog, Inventory Monitor, Cart Recovery) にルーティング',
     'Shopifyで注文発生→Webhook→agents/l2/inventory-monitor.ts が在庫更新。商品更新→Product Catalog同期。カート離脱→Cart Recovery起動',
     'npm run test -- agents/integration', '2', ['2-02'], 'CRITICAL', 'phase8'),

    ('8-02', '全47コンポーネント一斉起動テスト', 'agents/registration/agent-registration.ts (53.6KB)', 'PARTIAL',
     'agent-registration.ts の createAndBootstrap() を実行し全47体を起動。各エージェントのhealthCheck()がhealthyを返すことを確認。L0→L1→L2の起動順序が正しいことをログで検証',
     '/api/health が {"agents":47,"healthy":47,"pipelines":21}。全チーム(6)のリードが部下を認識。Commander.getSystemState() が全体ステータスを返す',
     'npm run test -- agents/registration/full-system-integration && npm run test -- agents/tests/phase18-e2e-integration', '2', ['7-08', '6-03', '4-04'], 'CRITICAL', 'phase8'),

    ('8-03', 'クロスパイプラインデータフロー検証 (15経路)', 'agents/pipelines/ (21パイプライン定義済み)', 'PARTIAL',
     '設計書v4の15クロスパイプラインデータフロー (P15→P01, P15→P05, P16→P05等) が全て動作することをE2Eテストで検証。agents/tests/cross-pipeline-e2e.test.ts を新規作成',
     '15経路すべてでデータが正しく流れる。P15 Go→P05 deploy, P16 CRITICAL→P05 emergency deploy, G-1→P07 analytics, I-1〜I-6→Self-Improvement学習 が動作',
     'npm run test -- agents/tests/cross-pipeline-e2e', '2', ['8-02'], 'CRITICAL', 'phase8'),

    ('8-04', 'Fly.ioフルデプロイ: 全47コンポーネント', '', 'MISSING',
     'fly.toml のvm.sizeをperformance-1xに変更 (47体同時稼働)。fly deploy。fly scale count 2 (冗長化)。全エージェントのヘルスチェック確認',
     'fly status でrunning。/api/health で47/47 healthy。Slack 4チャンネルに起動通知。メモリ使用量が制限内',
     'fly deploy && fly status && curl https://astromeda-agents.fly.dev/api/health', '1', ['8-03'], 'CRITICAL', 'phase8'),

    # ===== Phase 9: 出生 — ダッシュボード + Go-Live =====
    ('PH9', '【Phase 9】出生 — CEOダッシュボード + Go-Live', '', '', '', '', '', '', [], '', 'phase9'),

    ('9-01', 'CEOダッシュボード WebUI (27ウィジェット)', 'app/routes/admin.*.tsx (24ファイル既存)', 'PARTIAL',
     '既存admin画面にエージェント監視UIあり。GUI設計書の27ウィジェット (売上リアルタイム/目標ゲージ/コンバージョンファネル/エージェント状態6×4グリッド/SEOキーワードTop10/在庫アラート等) を追加。recharts + shadcn/ui 使用',
     'ブラウザでダッシュボードにアクセス→27ウィジェットが表示→リアルタイム更新(5秒間隔)→モバイルレスポンシブ',
     'npm run build && npm run dev でダッシュボードアクセス確認', '5', ['8-04'], 'HIGH', 'phase9'),

    ('9-02', '負荷テスト + パフォーマンスチューニング', '', 'MISSING',
     'k6 または Artillery で負荷テスト: 47エージェント同時リクエスト, 100並列パイプライン実行, Webhook 1000req/min シミュレーション。ボトルネック特定→メモリ/CPU最適化',
     '全テストシナリオでp95レスポンスタイム<2秒。メモリリークなし (1時間連続稼働でヒープ増加<10%)。Fly.io auto-scale動作',
     'npm run test:load', '2', ['8-04'], 'HIGH', 'phase9'),

    ('9-03', '災害復旧テスト', '', 'MISSING',
     'シナリオ: (1)Commander停止→自動復旧, (2)DB接続断→キュー蓄積→再接続→リプレイ, (3)Slack API断→ローカルログ→復旧時一括送信, (4)全エージェント同時再起動→データ損失なし',
     '全4シナリオでデータ損失ゼロ。復旧時間: Commander<30秒, DB<60秒, Slack<300秒。ランブックが自動生成',
     'npm run test -- agents/tests/disaster-recovery', '2', ['9-02'], 'CRITICAL', 'phase9'),

    ('9-04', 'EC本番Go-Live (v135→Production + エージェント稼働)', '', 'MISSING',
     'Step1: v135をProduction deploy (shopify hydrogen deploy --env production)。Step2: shop.mining-base.co.jp ドメインをHydrogenに切替。Step3: Shopify Webhook本番設定。Step4: 全エージェント本番モード有効化',
     'shop.mining-base.co.jp でHydrogenストアフロント表示。注文→Webhook→エージェント処理の一連が動作。72時間以内にロールバックなし',
     'curl -I https://shop.mining-base.co.jp && fly logs --app astromeda-agents', '1', ['9-03'], 'CRITICAL', 'phase9'),

    ('9-05', 'Go-Live後72時間ICU監視', '', 'MISSING',
     '24時間体制でSlack 4チャンネル監視。異常検知時の即時対応手順。KPI (売上/CVR/エラー率/応答時間) を1時間ごとにダッシュボードで確認',
     '72時間でCRITICALインシデントゼロ。売上が現行サイト比で95%以上維持。エラー率<0.1%',
     'fly logs --app astromeda-agents && curl https://astromeda-agents.fly.dev/api/health', '3', ['9-04'], 'CRITICAL', 'phase9'),

    # ===== Phase 10: 学習期 — 自律運営確立 =====
    ('PH10', '【Phase 10】学習期 — 自律運営の確立', '', '', '', '', '', '', [], '', 'phase10'),

    ('10-01', 'フィードバック学習ループ本稼働', '', 'MISSING',
     'Self-Improvement (#33) のプロンプト最適化サイクルを週次実行に設定。FeedbackAnalyzer の結果を Content Writer/SNS Manager/Ads Manager の system prompt に自動反映',
     '承認率が3ヶ月で50%→80%に向上。プロンプト変更履歴がDB + Slack #astromeda-品質 に記録',
     'npm run test -- agents/l2/self-improvement', '3', ['9-05'], 'HIGH', 'phase10'),

    ('10-02', '全エージェントKPI初回レビュー', '', 'MISSING',
     '設計書v7の各エージェントKPI (コンテンツ月200-300本, CVR改善, ROAS目標, 在庫回転率等) を計測。目標未達エージェントの原因分析 + 改善計画',
     'KPIダッシュボードに47体全エージェントのスコアカード表示。改善計画がSlack #astromeda-品質 にレポート',
     'curl https://astromeda-agents.fly.dev/api/kpi-report', '2', ['10-01'], 'HIGH', 'phase10'),

    ('10-03', 'Phase 2完了報告 + Phase 3計画', '', 'MISSING',
     '完了条件チェック: (1)47体稼働✓, (2)16PL動作✓, (3)Slack通知✓, (4)ダッシュボード✓, (5)EC連携✓, (6)CI/CD✓。Phase 3計画: マルチAI組合せ(Round 3-5), AI Red vs Blue自動化, Amazon/楽天マルチチャネル',
     '完了報告書がSlack + ダッシュボードに配信。Phase 3ロードマップが生成',
     '', '1', ['10-02'], 'HIGH', 'phase10'),
]

# ========== SHEET 1: 原子レベルガントチャート v2 ==========
ws = wb.active
ws.title = '原子レベルガントチャートv2'
ws.sheet_properties.tabColor = '0F3460'

ws.merge_cells('A1:K1')
ws['A1'] = 'Astromeda Phase 2 — 原子レベルガントチャート v2（実コード監査済み・曖昧さゼロ版）'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color=COLORS['dark_bg'])
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:K2')
ws['A2'] = '実装状態: ■完全実装 ■部分実装(TODO有) ■Stub(スケルトン) ■未実装 | 既存118ファイルと設計書v7-v12を突合監査済み'
ws['A2'].font = Font(name='Arial', size=9, color='666666', italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

headers = ['ID', 'タスク名', '既存ファイル', '実装状態', '正確な作業内容', '受入条件（曖昧さゼロ）', 'テストコマンド', '日数', '依存先', '優先度', '成熟段階']
col_widths = [7, 42, 38, 10, 55, 55, 38, 5, 18, 9, 9]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=4, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[4].height = 28

row = 5
for t in tasks:
    tid, name, existing, status, work, criteria, test, days, deps, pri, phase = t
    is_header = tid.startswith('PH')
    
    vals = [tid, name, existing, status, work, criteria, test, days, ', '.join(deps) if deps else '', pri, phase.replace('phase','Ph')]
    
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        
        if is_header:
            cell.font = phase_font
            cell.fill = PatternFill('solid', fgColor=COLORS.get(phase, 'FFFFFF'))
        else:
            # Status-based coloring for column 4
            if col_idx == 4:
                status_colors = {'COMPLETE': 'complete', 'PARTIAL': 'partial', 'STUB': 'stub', 'MISSING': 'missing', '要確認': 'partial'}
                cell.fill = PatternFill('solid', fgColor=COLORS.get(status_colors.get(val, ''), 'FFFFFF'))
                cell.font = Font(name='Arial', bold=True, size=8, color='333333')
            elif col_idx == 10:  # Priority
                if val == 'CRITICAL':
                    cell.font = critical_font
                elif val == 'HIGH':
                    cell.font = warn_font
                else:
                    cell.font = normal_font
                cell.fill = PatternFill('solid', fgColor=COLORS.get(phase, 'FFFFFF'))
            else:
                cell.font = normal_font
                cell.fill = PatternFill('solid', fgColor=COLORS.get(phase, 'FFFFFF'))
            
        cell.alignment = Alignment(horizontal='left' if col_idx in [2,3,5,6,7] else 'center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[row].height = 30 if is_header else 65
    row += 1

ws.freeze_panes = 'C5'

# ========== SHEET 2: 実コード監査レポート ==========
ws2 = wb.create_sheet('実コード監査レポート')
ws2.sheet_properties.tabColor = '66BB6A'

ws2.merge_cells('A1:F1')
ws2['A1'] = '実コード × 設計書v7-v12 突合監査レポート'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color=COLORS['dark_bg'])
ws2.row_dimensions[1].height = 30

audit_items = [
    ('agents/core/types.ts', 'COMPLETE', '162行', '全インターフェース定義済み (IAgent, IAgentBus, PipelineDefinition, CascadeCommand, AgentHealth, FeedbackRecord)', '設計書v7の型定義と100%一致', '追加作業不要'),
    ('agents/core/agent-bus.ts', 'COMPLETE', '307行', 'Pub/Sub + 優先キュー + Dead Letter Queue + Correlation ID', 'Slack通知hookの実接続が必要', 'Phase 1 タスク1-04'),
    ('agents/core/cascade-engine.ts', 'COMPLETE', '227行', 'L0→L1→L2カスケード + ロールバック + 500履歴', '設計書通り完全動作', '追加作業不要'),
    ('agents/l0/commander.ts', 'COMPLETE', '343行', 'Andon Cord + Graceful Shutdown + 9イベントハンドラ', '設計書v7のCommander仕様と一致', '追加作業不要'),
    ('agents/l1/base-lead.ts', 'COMPLETE', '378行', 'タスクキュー + AIBrain統合 + 6抽象メソッド', 'チーム名の設計書整合が必要', 'Phase 3 タスク3-01'),
    ('agents/l2/base-l2-agent.ts', 'COMPLETE', '231行', 'コマンドディスパッチ + エラー追跡 + 承認リクエスト', '設計書通り', '追加作業不要'),
    ('agents/providers/external-service-provider.ts', 'COMPLETE', '361行', 'IExternalServiceProvider + ProviderRegistry + StubProvider(181行)', '実APIは全てStub。TODO3箇所', 'Phase 5 タスク5-01〜5-05'),
    ('agents/providers/sns-providers.ts', 'STUB', '167行', 'X/Instagram/TikTok全てStubSNSProvider拡張', '3箇所TODO: 実API実装待ち', 'Phase 5 タスク5-01〜5-03'),
    ('agents/providers/ads-providers.ts', 'STUB', '243行', 'Google/Meta/LINE全てStubAdsProvider拡張。スタブデータはリアル値', '4箇所TODO: 実API実装待ち', 'Phase 5 タスク5-04〜5-05'),
    ('agents/data-collection/ga4-client.ts', 'PARTIAL', '507行', 'API呼出構造あり (analyticsdata.googleapis.com/v1beta)。JWT認証未実装', 'getAccessToken()がTODO', 'Phase 0 タスク0-08'),
    ('agents/data-collection/gsc-client.ts', 'PARTIAL', '386行', 'API呼出構造あり (searchconsole.googleapis.com/v3)。JWT認証未実装', 'getAccessToken()がTODO', 'Phase 0 タスク0-09'),
    ('agents/data-collection/competitor-scraper.ts', 'STUB', '412行', '7メーカー+ガジェットのStubデータ生成。実スクレイピング未実装', '全メソッドがStubデータ返却', 'Phase 4 タスク4-04'),
    ('agents/data-collection/ai-visibility-checker.ts', 'STUB', '281行', '確率ベースStub。4AIエンジン(ChatGPT/Gemini/Perplexity/Copilot)への実クエリ未実装', '全メソッドが確率ベース返却', 'Phase 4 タスク4-03'),
    ('agents/approval/approval-orchestrator.ts', 'COMPLETE', '410行', '7承認ポリシー + 自動承認エンジン + KPI追跡', 'Slack Block Kit UI未接続', 'Phase 6 タスク6-01'),
    ('agents/approval/feedback-analyzer.ts', 'COMPLETE', '387行', '10kフィードバック分析 + 学習サイクル + トレンド計算', 'Self-Improvement接続強化が必要', 'Phase 6 タスク6-03'),
    ('agents/pipelines/pipeline-engine.ts', 'COMPLETE', '668行', 'ステップ実行 + リトライ + ロールバック + 1000履歴', '設計書通り完全動作', '追加作業不要'),
    ('agents/pipelines/pipeline-definitions.ts', 'COMPLETE', '496行+', '21パイプライン定義済み (P01-P21)', 'P15/P16の詳細ステップ追加が必要', 'Phase 7 タスク7-07, 7-08'),
    ('agents/registry/agent-registry.ts', 'COMPLETE', '155行', 'CRUD + 能力検索 + 依存解決 + 循環検知', '設計書通り完全動作', '追加作業不要'),
    ('lib/validation/ (6ファイル)', 'MISSING', '0行', '統計エンジン/ラウンド実行/サンドボックス/攻撃エンジン/脆弱性マッパー/型定義 — 全て未実装', '設計書v5の検証基盤層が未着手', 'Phase 7 タスク7-01〜7-04'),
]

headers2 = ['ファイルパス', '状態', '行数', '実装内容', '設計書との差分', '対応タスク']
for col_idx, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 50
ws2.column_dimensions['E'].width = 35
ws2.column_dimensions['F'].width = 22
ws2.row_dimensions[3].height = 26

for row_idx, item in enumerate(audit_items, 4):
    for col_idx, val in enumerate(item, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = thin_border
        if col_idx == 2:
            sc = {'COMPLETE': 'complete', 'PARTIAL': 'partial', 'STUB': 'stub', 'MISSING': 'missing'}
            cell.fill = PatternFill('solid', fgColor=COLORS.get(sc.get(val, ''), 'FFFFFF'))
            cell.font = Font(name='Arial', bold=True, size=8)
    ws2.row_dimensions[row_idx].height = 45

# ========== SHEET 3: サマリー ==========
ws3 = wb.create_sheet('統計サマリー')
ws3.sheet_properties.tabColor = '6C63FF'

ws3.merge_cells('A1:D1')
ws3['A1'] = 'Phase 2 ガントチャート v2 — 統計サマリー'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color=COLORS['dark_bg'])

actual = [t for t in tasks if not t[0].startswith('PH')]
stats = [
    ('総タスク数', len(actual)),
    ('CRITICAL', sum(1 for t in actual if t[9] == 'CRITICAL')),
    ('HIGH', sum(1 for t in actual if t[9] == 'HIGH')),
    ('MEDIUM', sum(1 for t in actual if t[9] == 'MEDIUM')),
    ('', ''),
    ('既存コード状態', ''),
    ('COMPLETE (追加作業不要)', '7ファイル'),
    ('PARTIAL (JWT/名称/接続の修正)', '6ファイル'),
    ('STUB (実API差替え待ち)', '5ファイル'),
    ('MISSING (新規作成)', '6ファイル (lib/validation/ + Fly.io)'),
    ('', ''),
    ('新規作成が必要なファイル', ''),
    ('agents/core/ai-router.ts', 'Tier A-D モデルルーティング'),
    ('agents/core/gemini-provider.ts', 'Gemini Flash/Flash-Lite接続'),
    ('agents/core/fallback-manager.ts', '3リトライ+自動切替'),
    ('agents/core/notification-bus.ts', 'Slack 4チャンネル配信'),
    ('agents/lib/validation/types.ts', '検証共通型定義'),
    ('agents/lib/validation/statistical-engine.ts', 't検定/CI/Cohen d/CUSUM'),
    ('agents/lib/validation/round-executor.ts', 'ラウンド反復制御'),
    ('agents/lib/validation/sandbox-manager.ts', 'Fly.ioサンドボックス管理'),
    ('agents/lib/validation/attack-engine.ts', '攻撃実行エンジン'),
    ('agents/lib/validation/vulnerability-mapper.ts', 'CVSS v4.0マッピング'),
    ('agents/l2/ai-capability-validator.ts', '#39 Validator (7ステップ)'),
    ('agents/l2/ai-security-auditor.ts', '#40 Auditor (6ステップ)'),
    ('agents/approval/slack-approval-bot.ts', 'Block Kit UI承認ボット'),
    ('agents/approval/content-approval-gate.ts', '承認ゲート'),
    ('lib/databases/schema.sql', 'PostgreSQL 6テーブル'),
    ('lib/databases/migrate.ts', 'マイグレーションスクリプト'),
    ('fly.toml', 'Fly.ioデプロイ設定'),
    ('.github/workflows/deploy-agents.yml', 'CI/CDワークフロー'),
]

for col_idx, h in enumerate(['項目', '値'], 1):
    cell = ws3.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

ws3.column_dimensions['A'].width = 45
ws3.column_dimensions['B'].width = 40

for row_idx, (k, v) in enumerate(stats, 4):
    ws3.cell(row=row_idx, column=1, value=k).font = Font(name='Arial', bold=True if not k.startswith('agents') and not k.startswith('lib') and not k.startswith('fly') and not k.startswith('.github') else False, size=9)
    ws3.cell(row=row_idx, column=2, value=v).font = normal_font
    ws3.cell(row=row_idx, column=1).border = thin_border
    ws3.cell(row=row_idx, column=2).border = thin_border

# Save
out = '/sessions/nifty-festive-ramanujan/mnt/市場調査/Astromeda_Phase2_原子レベルガントチャートv2_実コード監査版.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Tasks: {len(actual)}, CRITICAL: {sum(1 for t in actual if t[9]=="CRITICAL")}, HIGH: {sum(1 for t in actual if t[9]=="HIGH")}')
